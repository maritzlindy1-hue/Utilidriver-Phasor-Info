from __future__ import annotations

import io
import math
import os
import time
import uuid
import xml.etree.ElementTree as ET
from dataclasses import dataclass, asdict
from typing import Any

import requests
from flask import Flask, redirect, render_template_string, request, send_file, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

API_URL = os.getenv("UTILIDRIVER_API_URL", "http://197.189.218.35/utilidriver/api/orders/")
REQUEST_TIMEOUT_SECONDS = int(os.getenv("REQUEST_TIMEOUT_SECONDS", "20"))
POLL_INTERVAL_SECONDS = float(os.getenv("POLL_INTERVAL_SECONDS", "2"))
POLL_TIMEOUT_SECONDS = float(os.getenv("POLL_TIMEOUT_SECONDS", "120"))

app = Flask(__name__)
RESULT_CACHE: dict[str, dict[str, Any]] = {}

REGISTER_LABELS = {
    "1.1.21.7.0.255": "L1 Active Power", "1.1.22.7.0.255": "L1 Reactive Power",
    "1.1.31.7.0.255": "L1 Current", "1.1.32.7.0.255": "L1 Voltage", "1.1.33.7.0.255": "L1 Power Factor / Angle",
    "1.1.41.7.0.255": "L2 Active Power", "1.1.42.7.0.255": "L2 Reactive Power",
    "1.1.51.7.0.255": "L2 Current", "1.1.52.7.0.255": "L2 Voltage", "1.1.53.7.0.255": "L2 Power Factor / Angle",
    "1.1.61.7.0.255": "L3 Active Power", "1.1.62.7.0.255": "L3 Reactive Power",
    "1.1.71.7.0.255": "L3 Current", "1.1.72.7.0.255": "L3 Voltage", "1.1.73.7.0.255": "L3 Power Factor / Angle",
}

PHASES = {
    "L1": {"v": "1.1.32.7.0.255", "i": "1.1.31.7.0.255", "pf": "1.1.33.7.0.255", "base_angle": 0},
    "L2": {"v": "1.1.52.7.0.255", "i": "1.1.51.7.0.255", "pf": "1.1.53.7.0.255", "base_angle": -120},
    "L3": {"v": "1.1.72.7.0.255", "i": "1.1.71.7.0.255", "pf": "1.1.73.7.0.255", "base_angle": 120},
}

PAGE = """
<!doctype html><html><head><meta charset="utf-8"><title>UtiliDriver Phasor / CT Info</title>
<style>
body{font-family:Arial,sans-serif;background:#f4f6f8;margin:0;padding:30px;color:#1f2937}.card{max-width:1180px;margin:auto;background:white;padding:24px;border-radius:14px;box-shadow:0 8px 30px #0001}input,button,.btn{font-size:16px;padding:12px;border-radius:8px;border:1px solid #cbd5e1}button,.btn{background:#111827;color:white;cursor:pointer;text-decoration:none;display:inline-block;margin:4px 4px 4px 0}.btn.blue{background:#2563eb}.btn.green{background:#047857}.ok{background:#ecfdf5;border-left:5px solid #10b981;padding:12px}.bad{background:#fef2f2;border-left:5px solid #ef4444;padding:12px}table{border-collapse:collapse;width:100%;margin-top:18px}th,td{border-bottom:1px solid #e5e7eb;padding:10px;text-align:left}th{background:#f9fafb}.small{color:#64748b;font-size:13px}.urls{word-break:break-all;background:#f8fafc;padding:10px;border-radius:8px}.grid{display:grid;grid-template-columns:1fr 1fr;gap:20px}.phasor{max-width:100%;border:1px solid #e5e7eb;border-radius:12px;background:white}@media(max-width:900px){.grid{grid-template-columns:1fr}}
</style></head><body><div class="card">
<h1>UtiliDriver Phasor / CT Info</h1>
<p class="small">Step 1 gets readings only. It sends only one POST order. Phasor/PDF/Excel run only after the readings are already loaded.</p>
<form method="post"><input name="meter_number" placeholder="e.g. 12345678" maxlength="8" pattern="[0-9]{8}" required><button type="submit">Get Readings Once</button></form>
{% if result %}<h2>Result</h2><div class="{{ 'ok' if result.ok else 'bad' }}">{{ result.message }}</div>
{% if result.meter_dec %}<p><b>Meter:</b> {{ result.meter_dec }} {% if result.meter_hex %} | <b>Hex:</b> {{ result.meter_hex }}{% endif %}</p>{% endif %}
{% if result.order_url %}<div class="urls"><b>Order URL:</b> {{ result.order_url }}<br><b>Status URL:</b> {{ result.status_url }}<br><b>Completed URL:</b> {{ result.completed_url }}</div>{% endif %}
{% if result.ct_ratio %}<p><b>CT Ratio:</b> {{ result.ct_ratio }} {% if result.ct_fraction %} | <b>Fraction:</b> {{ result.ct_fraction }}{% endif %}</p>{% endif %}
{% if result.ok and result.result_id %}<p><a class="btn green" href="{{ url_for('show_phasor', result_id=result.result_id) }}">Draw Phasor Diagram</a><a class="btn blue" href="{{ url_for('download_excel', result_id=result.result_id) }}">Download Excel</a><a class="btn blue" href="{{ url_for('download_pdf', result_id=result.result_id) }}">Download PDF</a></p>{% endif %}
{% if phasor_svg %}<div class="grid"><div><h3>Approximate Phasor Diagram</h3>{{ phasor_svg|safe }}<p class="small">Voltage angles are assumed at L1 0°, L2 -120°, L3 +120°. Current angle is estimated from PF using acos(PF). Reverse lag/lead sign later if the meter convention proves opposite.</p></div><div><h3>Phase Summary</h3><table><tr><th>Phase</th><th>Voltage</th><th>Current</th><th>PF</th><th>V Angle</th><th>I Angle</th></tr>{% for p in phase_summary %}<tr><td>{{p.phase}}</td><td>{{p.voltage}}</td><td>{{p.current}}</td><td>{{p.pf}}</td><td>{{p.v_angle}}</td><td>{{p.i_angle}}</td></tr>{% endfor %}</table></div></div>{% endif %}
{% if result.rows %}<h3>Registers</h3><table><tr><th>Register</th><th>Description</th><th>Unit</th><th>Scale</th><th>Value</th></tr>{% for row in result.rows %}<tr><td>{{ row.id }}</td><td>{{ row.description }}</td><td>{{ row.unit }}</td><td>{{ row.scale }}</td><td>{{ row.value }}</td></tr>{% endfor %}</table>{% endif %}
{% endif %}</div></body></html>
"""

@dataclass
class OrderResult:
    ok: bool
    message: str
    meter_dec: str | None = None
    meter_hex: str | None = None
    order_url: str | None = None
    status_url: str | None = None
    completed_url: str | None = None
    ct_ratio: str | None = None
    ct_fraction: str | None = None
    rows: list[dict[str, str]] | None = None
    result_id: str | None = None


def validate_meter_number(meter_dec: str) -> str:
    meter_dec = meter_dec.strip()
    if len(meter_dec) != 8 or not meter_dec.isdigit():
        raise ValueError("Meter number must be exactly 8 digits.")
    return meter_dec


def decimal_to_hex_8(meter_dec: str) -> str:
    return format(int(meter_dec), "08X")


def build_xml_body(meter_hex: str) -> str:
    registers = "\n".join(f"    <RegisterCommand action='read'><Register id='{reg_id}' /></RegisterCommand>" for reg_id in REGISTER_LABELS)
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<Order priority='High'>
  <Subjects><Meters><Meter ref="http://197.189.218.35/utilidriver/api/meters/4B414D000000{meter_hex}/" /></Meters></Subjects>
  <Commands>
    <TransformerRatioConfigurationCommand action='read' />
{registers}
  </Commands>
</Order>""".strip()


def send_order_once(xml_body: str) -> str:
    headers = {"Content-Type": "application/xml", "Accept": "application/xml"}
    response = requests.post(API_URL, data=xml_body.encode("utf-8"), headers=headers, timeout=REQUEST_TIMEOUT_SECONDS)
    response.raise_for_status()
    order_url = response.headers.get("Location")
    if not order_url:
        raise RuntimeError("Order posted, but no Location header was returned.")
    return order_url


def wait_for_order(status_url: str) -> bool:
    end_time = time.time() + POLL_TIMEOUT_SECONDS
    while time.time() < end_time:
        response = requests.get(status_url, timeout=REQUEST_TIMEOUT_SECONDS)
        response.raise_for_status()
        root = ET.fromstring(response.text.strip())
        waiting_text = root.findtext("WaitingCommandCount")
        if waiting_text is None:
            raise RuntimeError("WaitingCommandCount was not found in status XML.")
        if int(waiting_text.strip()) == 0:
            return True
        time.sleep(POLL_INTERVAL_SECONDS)
    return False


def parse_completed_results(completed_url: str):
    response = requests.get(completed_url, timeout=REQUEST_TIMEOUT_SECONDS)
    response.raise_for_status()
    root = ET.fromstring(response.text)
    rows = []
    for reg in root.findall(".//Register"):
        reg_id = (reg.get("id") or "").strip()
        if reg_id:
            rows.append({
                "id": reg_id,
                "description": REGISTER_LABELS.get(reg_id, ""),
                "unit": (reg.findtext("Unit") or "").strip(),
                "scale": (reg.findtext("Scale") or "").strip(),
                "value": (reg.findtext("Value") or "").strip(),
            })
    rows.sort(key=lambda row: row["id"])
    ct_ratio = (root.findtext(".//CurrentTransformerRatio") or "").strip() or None
    ct_fraction = None
    if ct_ratio:
        try:
            ct_fraction = f"{int(ct_ratio) * 5}/5"
        except ValueError:
            pass
    return rows, ct_ratio, ct_fraction


def run_meter_order(meter_dec: str) -> OrderResult:
    try:
        meter_dec = validate_meter_number(meter_dec)
        meter_hex = decimal_to_hex_8(meter_dec)
        order_url = send_order_once(build_xml_body(meter_hex))  # exactly one POST only
        status_url = f"{order_url.rstrip('/')}/status/"
        completed_url = f"{order_url.rstrip('/')}/completed/"
        if not wait_for_order(status_url):
            return OrderResult(False, "Timed out. No second POST was sent.", meter_dec, meter_hex, order_url, status_url, completed_url)
        rows, ct_ratio, ct_fraction = parse_completed_results(completed_url)
        result = OrderResult(True, "Order completed successfully. Readings loaded. You can now draw the phasor or download Excel/PDF.", meter_dec, meter_hex, order_url, status_url, completed_url, ct_ratio, ct_fraction, rows)
        result.result_id = uuid.uuid4().hex
        RESULT_CACHE[result.result_id] = asdict(result)
        return result
    except Exception as exc:
        return OrderResult(False, f"Stopped after one attempt: {exc}", meter_dec if meter_dec else None)


def _float_value(row_map: dict[str, dict[str, str]], reg_id: str) -> float | None:
    try:
        return float(row_map[reg_id]["value"])
    except Exception:
        return None


def build_phase_summary(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    row_map = {r["id"]: r for r in rows}
    summary = []
    for phase, info in PHASES.items():
        v = _float_value(row_map, info["v"])
        i = _float_value(row_map, info["i"])
        pf = _float_value(row_map, info["pf"])
        base = float(info["base_angle"])
        pf_clamped = max(-1, min(1, pf if pf is not None else 1))
        phi = math.degrees(math.acos(abs(pf_clamped))) if pf is not None else 0
        current_angle = base - phi
        summary.append({
            "phase": phase,
            "voltage": "" if v is None else f"{v:g} V",
            "current": "" if i is None else f"{i:g} A",
            "pf": "" if pf is None else f"{pf:g}",
            "v_angle": f"{base:g}°",
            "i_angle": f"{current_angle:.1f}°",
            "i_angle_raw": current_angle,
            "v_angle_raw": base,
        })
    return summary


def phasor_svg(rows: list[dict[str, str]]) -> str:
    summary = build_phase_summary(rows)
    cx, cy, radius = 250, 250, 190
    parts = [
        '<svg class="phasor" width="520" height="520" viewBox="0 0 500 500" xmlns="http://www.w3.org/2000/svg">',
        '<defs><marker id="arrow" markerWidth="8" markerHeight="8" refX="7" refY="3" orient="auto"><path d="M0,0 L0,6 L7,3 z" fill="currentColor" /></marker></defs>',
        '<rect width="500" height="500" fill="white"/>',
        f'<circle cx="{cx}" cy="{cy}" r="{radius}" fill="none" stroke="#d1d5db"/>',
        f'<line x1="40" y1="{cy}" x2="460" y2="{cy}" stroke="#e5e7eb"/>',
        f'<line x1="{cx}" y1="40" x2="{cx}" y2="460" stroke="#e5e7eb"/>',
        '<text x="250" y="24" text-anchor="middle" font-size="18" font-weight="700">Approximate Phasor Diagram</text>',
    ]
    colors_list = {"L1":"#dc2626", "L2":"#2563eb", "L3":"#16a34a"}
    for p in summary:
        phase = p["phase"]
        col = colors_list.get(phase, "#111827")
        for label, angle, length, dash in [(f"{phase} V", p["v_angle_raw"], 1.0, ""), (f"{phase} I", p["i_angle_raw"], 0.72, "6,5")]:
            rad = math.radians(angle)
            x = cx + radius * length * math.cos(rad)
            y = cy - radius * length * math.sin(rad)
            lx = cx + (radius * length + 22) * math.cos(rad)
            ly = cy - (radius * length + 22) * math.sin(rad)
            dash_attr = f' stroke-dasharray="{dash}"' if dash else ""
            parts.append(f'<line x1="{cx}" y1="{cy}" x2="{x:.1f}" y2="{y:.1f}" stroke="{col}" stroke-width="3" marker-end="url(#arrow)"{dash_attr}/>' )
            parts.append(f'<text x="{lx:.1f}" y="{ly:.1f}" text-anchor="middle" font-size="13" fill="{col}" font-weight="700">{label}</text>')
    parts.append('</svg>')
    return "".join(parts)


@app.route("/", methods=["GET", "POST"])
def index() -> Any:
    result = run_meter_order(request.form.get("meter_number", "")) if request.method == "POST" else None
    return render_template_string(PAGE, result=result, phasor_svg=None, phase_summary=[])


@app.route("/phasor/<result_id>")
def show_phasor(result_id: str):
    data = RESULT_CACHE.get(result_id)
    if not data:
        return redirect(url_for("index"))
    return render_template_string(PAGE, result=data, phasor_svg=phasor_svg(data.get("rows") or []), phase_summary=build_phase_summary(data.get("rows") or []))


@app.route("/download/excel/<result_id>")
def download_excel(result_id: str):
    data = RESULT_CACHE.get(result_id)
    if not data:
        return redirect(url_for("index"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Meter Results"
    ws.append(["UtiliDriver Phasor / CT Info"])
    ws["A1"].font = Font(size=16, bold=True)
    meta = [["Meter", data.get("meter_dec")], ["Hex", data.get("meter_hex")], ["CT Ratio", data.get("ct_ratio")], ["CT Fraction", data.get("ct_fraction")], ["Order URL", data.get("order_url")], ["Completed URL", data.get("completed_url")]]
    for row in meta:
        ws.append(row)
    ws.append([])
    ws.append(["Phase", "Voltage", "Current", "PF", "Voltage Angle", "Current Angle"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="E5E7EB")
    for p in build_phase_summary(data.get("rows") or []):
        ws.append([p["phase"], p["voltage"], p["current"], p["pf"], p["v_angle"], p["i_angle"]])
    ws.append([])
    ws.append(["Register", "Description", "Unit", "Scale", "Value"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="E5E7EB")
    for r in data.get("rows") or []:
        ws.append([r["id"], r["description"], r["unit"], r["scale"], r["value"]])
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"meter_results_{data.get('meter_dec')}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/download/pdf/<result_id>")
def download_pdf(result_id: str):
    data = RESULT_CACHE.get(result_id)
    if not data:
        return redirect(url_for("index"))
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    story = [Paragraph("UtiliDriver Phasor / CT Info", styles["Title"]), Spacer(1, 0.2*cm)]
    story.append(Paragraph(f"Meter: {data.get('meter_dec')} &nbsp;&nbsp; Hex: {data.get('meter_hex')} &nbsp;&nbsp; CT: {data.get('ct_fraction') or data.get('ct_ratio')}", styles["Normal"]))
    story.append(Spacer(1, 0.3*cm))
    phase_rows = [["Phase", "Voltage", "Current", "PF", "V Angle", "I Angle"]] + [[p["phase"], p["voltage"], p["current"], p["pf"], p["v_angle"], p["i_angle"]] for p in build_phase_summary(data.get("rows") or [])]
    pt = Table(phase_rows)
    pt.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.lightgrey), ("GRID", (0,0), (-1,-1), 0.25, colors.grey), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold")]))
    story += [pt, Spacer(1, 0.3*cm)]
    rows = [["Register", "Description", "Unit", "Scale", "Value"]] + [[r["id"], r["description"], r["unit"], r["scale"], r["value"]] for r in data.get("rows") or []]
    rt = Table(rows)
    rt.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.lightgrey), ("GRID", (0,0), (-1,-1), 0.25, colors.grey), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"), ("FONTSIZE", (0,0), (-1,-1), 8)]))
    story.append(rt)
    doc.build(story)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"meter_results_{data.get('meter_dec')}.pdf", mimetype="application/pdf")


@app.route("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=False)
